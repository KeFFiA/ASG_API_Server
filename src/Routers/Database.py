import random
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Request, BackgroundTasks
from fastapi.responses import FileResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from pydantic.v1 import EmailStr
from sqlalchemy import select, inspect

from Config import RESPONSES_PATH, PA_APP_URL, CUSTOM_EXCEL_LEASE_HEADERS_ORDER
from Database.Models import Lease_Output
from Schemas import JsonFileSchema
from Schemas.Enums import service
from Utils import remove_file

router = APIRouter(
    prefix="/database",
    tags=["Database"]
)


@router.get('/{type}')
async def get_db(type: str, request: Request, background_tasks: BackgroundTasks):
    if type.lower() == 'lease_agr':
        main_db = await request.state.db.get_db("main")
        result = await main_db.execute(
            select(Lease_Output)
            .order_by(Lease_Output.id.asc())
        )

        rows = result.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Lease Agreements"

        mapper = inspect(Lease_Output)
        headers = [column.key for column in mapper.attrs]

        formatted_headers = [header.replace('_', ' ').title() for header in headers]

        button_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        button_font = Font(color="FFFFFF", bold=True, size=12)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(color="1F497D", bold=True, size=11)
        data_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(formatted_headers))
        cell = ws.cell(row=1, column=1, value="← Back to Application")
        cell.hyperlink = PA_APP_URL
        cell.font = button_font
        cell.fill = button_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

        # btn height
        ws.row_dimensions[1].height = 30

        # tbl headers
        for col_idx, header in enumerate(formatted_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header if header.lower() not in ["msn"] else header.upper())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # headers height
        ws.row_dimensions[2].height = 25

        # table data
        for row_idx, row in enumerate(rows, start=3):
            for col_idx, col_key in enumerate(headers, start=1):
                value = getattr(row, col_key, None)
                if value is None:
                    cell_value = ""
                elif isinstance(value, (datetime, date)):
                    cell_value = value.isoformat()
                else:
                    cell_value = str(value)

                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = data_font
                cell.border = border

                # white/gray per line
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for col_idx, col_name in enumerate(formatted_headers, start=1):
            max_len = len(col_name)
            for row_idx in range(3, len(rows) + 3):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_len = min(max(max_len, len(str(cell_value))), 50)
            adjusted_width = max_len + 3
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        link_row = len(rows) + 3
        ws.merge_cells(start_row=link_row, start_column=1, end_row=link_row, end_column=len(formatted_headers))

        cell = ws.cell(row=link_row, column=1, value="← Back to Application")
        cell.hyperlink = PA_APP_URL
        cell.font = button_font
        cell.fill = button_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

        # bottom btn height
        ws.row_dimensions[link_row].height = 30

        # filters for headers
        ws.auto_filter.ref = f"A2:{get_column_letter(len(formatted_headers))}{len(rows) + 2}"

        # freeze
        ws.freeze_panes = "A3"

        filename_xl = "Lease_Agreements.xlsx"
        filepath_xl = RESPONSES_PATH / filename_xl
        wb.save(filepath_xl)
    else:
        filename_xl = "Lease_Agreements.xlsx"

    data = JsonFileSchema(
        type=type,
        user_email=EmailStr("integrator@ai12.com"),
        filename=filename_xl
    )

    filename = f"{random.randint(10000, 99999)}.json"
    filepath = Path(RESPONSES_PATH / filename)
    filepath.write_text(data.model_dump_json(indent=4), encoding="utf-8")

    background_tasks.add_task(remove_file, str(filepath))

    return FileResponse(
        filepath,
        media_type="application/json",
        filename=filename,
        background=background_tasks,
    )


